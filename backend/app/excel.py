from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import ExportContext

RED = "D71920"
BLACK = "111111"
WHITE = "FFFFFF"
SOFT = "F5F6F8"
LINE = "D8DDE4"


def safe_sheet_name(value: str) -> str:
    cleaned = "".join(ch for ch in str(value or "Report") if ch not in r"[]:*?/\\")[:31]
    return cleaned or "Report"


def is_money_header(header: str) -> bool:
    text = str(header or "").lower()
    return any(token in text for token in ["amount", "total", "balance", "debit", "credit", "income", "cost", "profit", "loss", "inflow", "outflow", "net", "paid"])


def parse_money(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None
    text = str(value).replace("BWP", "").replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        return None


def build_report_excel(payload: dict[str, Any], context: ExportContext | None = None) -> bytes:
    report = payload.get("report") if isinstance(payload.get("report"), dict) else payload
    title = report.get("title") or report.get("reportName") or "Civil-Gineer Masta Report"
    headers = report.get("headers") or []
    rows = report.get("rows") or []
    filters = report.get("filters") or {}
    generated_at = report.get("generatedAt") or datetime.now().strftime("%Y-%m-%d %H:%M")
    company = context.settings.companyProfile if context else None
    company_name = (company.name if company else "") or "Civil-Gineer Masta Proprietary Limited"

    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_name(title)
    ws.sheet_view.showGridLines = False

    column_count = max(1, len(headers))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    ws["A1"] = company_name
    ws["A1"].font = Font(bold=True, size=15, color=BLACK)
    ws["A2"] = title
    ws["A2"].font = Font(bold=True, size=13, color=RED)
    ws["A3"] = f"Generated: {generated_at}"
    ws["A3"].font = Font(size=9, color="5F6672")

    row_index = 4
    if filters:
        for key, value in filters.items():
            ws.cell(row=row_index, column=1, value=str(key).replace("_", " ").title()).font = Font(bold=True)
            ws.cell(row=row_index, column=2, value=value)
            row_index += 1
    row_index += 1

    header_row = row_index
    for col_index, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_index, value=header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLACK)
        cell.alignment = Alignment(horizontal="center")

    money_columns = {index for index, header in enumerate(headers, 1) if is_money_header(header)}
    for source_row in rows:
        row_index += 1
        for col_index, value in enumerate(source_row, 1):
            cell = ws.cell(row=row_index, column=col_index)
            if col_index in money_columns:
                number = parse_money(value)
                cell.value = number if number is not None else value
                if number is not None:
                    cell.number_format = '"BWP" #,##0.00;[Red]-"BWP" #,##0.00'
            else:
                cell.value = value

    total_row = row_index + 1
    has_totals = row_index >= header_row + 1 and bool(money_columns)
    if has_totals:
        ws.cell(row=total_row, column=1, value="Totals").font = Font(bold=True)
        for col_index in money_columns:
            letter = get_column_letter(col_index)
            cell = ws.cell(row=total_row, column=col_index, value=f"=SUM({letter}{header_row + 1}:{letter}{row_index})")
            cell.font = Font(bold=True)
            cell.number_format = '"BWP" #,##0.00;[Red]-"BWP" #,##0.00'

    table_end_row = total_row if has_totals else row_index
    if headers and table_end_row > header_row:
        ref = f"A{header_row}:{get_column_letter(column_count)}{table_end_row}"
        excel_table = Table(displayName="CGMReportTable", ref=ref)
        excel_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(excel_table)

    ws.freeze_panes = f"A{header_row + 1}"
    thin = Side(style="thin", color=LINE)
    for row in ws.iter_rows(min_row=header_row, max_row=max(table_end_row, header_row), min_col=1, max_col=column_count):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_index in range(1, column_count + 1):
        values = [ws.cell(row=row, column=col_index).value for row in range(1, max(table_end_row, header_row) + 1)]
        width = min(42, max(12, max(len(str(value or "")) for value in values) + 2))
        ws.column_dimensions[get_column_letter(col_index)].width = width

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
