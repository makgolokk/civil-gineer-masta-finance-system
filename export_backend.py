from __future__ import annotations

import json
import os
from html import escape
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    KeepTogether,
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

HOST = os.getenv("HOST", "127.0.0.1")
PORT = int(os.getenv("PORT", "8765"))
CURRENCY = "BWP"

COMPANY = {
    "name": "Civil-Gineer Masta (Pty) Ltd",
    "subtitle": "BUILDING THE FUTURE, MASTERING THE PRESENT",
    "address": "Plot 31848, Gaborone North, Gaborone, Botswana",
    "phone": "+267 71839730 / +267 77008234",
    "email": "makgolokk@outlook.com",
    "terms": "Payment due strictly as per agreed milestones or due date stated on the document.",
    "notes": "Council fees, printing, copying, plotting, and any additional services outside the agreed scope may be billed separately.",
    "logoPath": "assets/logo.png",
    "bank": "FNB / RMB",
    "accountHolder": "Civil-Gineer Masta (Pty) Ltd",
    "accountType": "Business Cheque Account",
    "accountNumber": "63119613734",
    "branchName": "Gaborone Industrial",
    "branchCode": "283567",
    "preparedBy": "Kelesitse K. Makgolo",
    "approvedBy": "",
}

RED = "D71920"
BLACK = "111111"
WHITE = "FFFFFF"
LIGHT = "F5F6F8"
INK = "1F2328"
MUTED = "5F6672"
BORDER = "D8DDE4"
PAPER_TINT = "FAFAFB"


def money(value):
    return f"{CURRENCY} {float(value or 0):,.2f}"


def line_total(item):
    return float(item.get("qty") or 0) * float(item.get("rate") or 0)


def document_total(document):
    subtotal = sum(line_total(item) for item in document.get("items", []))
    discount = float(document.get("discount") or 0)
    tax = float(document.get("taxAmount") or max(0, subtotal - discount) * (float(document.get("taxRate") or 0) / 100))
    return max(0, subtotal - discount + tax)


def company_from_data(data):
    settings = (data or {}).get("settings", {})
    profile = settings.get("companyProfile", {})
    bank = profile.get("bankingDetails", {})
    company = dict(COMPANY)
    company.update({
        "name": profile.get("name") or company["name"],
        "subtitle": profile.get("letterhead") or company["subtitle"],
        "address": profile.get("address") or company["address"],
        "phone": " / ".join([v for v in [profile.get("phone"), profile.get("alternatePhone")] if v]) or company["phone"],
        "email": profile.get("email") or company["email"],
        "terms": profile.get("defaultTerms") or company["terms"],
        "notes": profile.get("defaultNotes") or company["notes"],
        "logoPath": profile.get("logoPath") or company["logoPath"],
        "bank": bank.get("bank") or company["bank"],
        "accountHolder": bank.get("accountHolder") or company["accountHolder"],
        "accountType": bank.get("accountType") or company["accountType"],
        "accountNumber": bank.get("accountNumber") or company["accountNumber"],
        "branchName": bank.get("branchName") or company["branchName"],
        "branchCode": bank.get("branchCode") or company["branchCode"],
        "preparedBy": profile.get("preparedBy") or company["preparedBy"],
        "approvedBy": profile.get("approvedBy") or company["approvedBy"],
    })
    return company


def invoice_paid(data, invoice_id):
    return sum(float(payment.get("amount") or 0) for payment in data.get("payments", []) if payment.get("invoiceId") == invoice_id)


def invoice_status(data, invoice):
    total = document_total(invoice)
    paid = invoice_paid(data, invoice.get("id"))
    if paid <= 0:
        return "Unpaid"
    if paid + 0.01 >= total:
        return "Paid"
    return "Partial"


def find_by_id(items, item_id):
    return next((item for item in items if item.get("id") == item_id), None)


def client_name(data, client_id):
    client = find_by_id(data.get("clients", []), client_id) or {}
    return client.get("name", "")


def service_name(data, service_id):
    service = find_by_id(data.get("services", []), service_id) or {}
    return service.get("name", "General")


def supplier_name(data, supplier_id):
    supplier = find_by_id(data.get("suppliers", []), supplier_id) or {}
    return supplier.get("name", "")


def account_name(data, account_id):
    account = find_by_id(data.get("accounts", []), account_id) or {}
    if account:
        return f"{account.get('code', '')} - {account.get('name', '')}".strip(" -")
    return account_id or ""


def supplier_bill_number(data, bill_id):
    bill = find_by_id(data.get("supplierBills", []), bill_id) or {}
    return bill.get("number", bill_id or "")


def quotation_client_name(data, quote):
    if quote.get("clientId"):
        return client_name(data, quote.get("clientId"))
    snapshot = quote.get("clientSnapshot") or {}
    return snapshot.get("name", "Unlinked prospect")


def workbook_project_label(data, item):
    project = find_by_id(data.get("projects", []), item.get("projectId")) or {}
    code = project.get("code") or item.get("projectCode") or ""
    name = project.get("name") or item.get("projectName") or ""
    if code and name:
        return f"{code} - {name}"
    return code or name or ""


def month_key(value):
    return str(value or "")[:7]


def current_month_summary(data):
    month = datetime.now().strftime("%Y-%m")
    invoices = [item for item in data.get("invoices", []) if month_key(item.get("date")) == month]
    payments = [item for item in data.get("payments", []) if month_key(item.get("date")) == month]
    expenses = [item for item in data.get("expenses", []) if month_key(item.get("date")) == month]
    invoiced = sum(document_total(item) for item in invoices)
    received = sum(float(item.get("amount") or 0) for item in payments)
    spent = sum(float(item.get("amount") or 0) for item in expenses)
    return {
        "month": month,
        "invoiced": invoiced,
        "received": received,
        "expenses": spent,
        "net": received - spent,
    }


def stylesheet():
    styles = getSampleStyleSheet()
    styles["Normal"].fontName = "Helvetica"
    styles["Normal"].fontSize = 9
    styles["Normal"].leading = 12
    styles["Normal"].textColor = colors.HexColor(f"#{INK}")
    styles.add(
        ParagraphStyle(
            name="SmallMuted",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor(f"#{MUTED}"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="DocTitle",
            parent=styles["Heading1"],
            fontSize=23,
            leading=26,
            textColor=colors.HexColor(f"#{RED}"),
            spaceAfter=6,
        )
    )
    styles.add(
        ParagraphStyle(
            name="DocKicker",
            parent=styles["Normal"],
            fontSize=7.5,
            leading=9,
            textColor=colors.HexColor(f"#{RED}"),
            fontName="Helvetica-Bold",
            uppercase=True,
        )
    )
    styles.add(
        ParagraphStyle(
            name="BoxTitle",
            parent=styles["Normal"],
            fontSize=8.5,
            leading=10,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor(f"#{BLACK}"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="SectionTitle",
            parent=styles["Normal"],
            fontSize=10,
            leading=12,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor(f"#{BLACK}"),
            spaceAfter=4,
        )
    )
    styles.add(
        ParagraphStyle(
            name="DocNumber",
            parent=styles["Normal"],
            fontSize=9,
            leading=11,
            textColor=colors.HexColor(f"#{MUTED}"),
            alignment=2,
        )
    )
    styles.add(
        ParagraphStyle(
            name="AmountHero",
            parent=styles["Normal"],
            fontSize=26,
            leading=30,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor(f"#{RED}"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="Footer",
            parent=styles["SmallMuted"],
            alignment=1,
        )
    )
    return styles


def p(text, style):
    return Paragraph(escape(str(text or "")).replace("\n", "<br/>"), style)


def rich(text, style):
    return Paragraph(str(text or ""), style)


def brand_header(title, number, styles, data=None):
    company = company_from_data(data)
    try:
        logo_path = Path(company["logoPath"])
        doc_logo = logo_path.with_name("logo-doc.png") if logo_path.name == "logo.png" else logo_path
        logo = Image(str(doc_logo if doc_logo.exists() else logo_path), width=48 * mm, height=28 * mm)
    except Exception:
        logo = rich(f"<b>{escape(company['name'])}</b>", styles["Normal"])
    title_block = [
        rich(f"<font size='7' color='#{MUTED}'><b>CLIENT DOCUMENT</b></font>", styles["DocNumber"]),
        rich(f"<font size='26' color='#{RED}'><b>{escape(title.upper())}</b></font>", styles["Normal"]),
        rich(f"<font color='#{BLACK}'><b>No. {escape(str(number or 'Draft'))}</b></font>", styles["DocNumber"]),
    ]
    contact = f"{company['address']}<br/>{company['phone']}<br/>{company['email']}"
    return Table(
        [
            [
                logo,
                rich(f"<font size='12'><b>{escape(company['name'])}</b></font><br/><font color='#{RED}'><b>{escape(company['subtitle'])}</b></font><br/><font size='8' color='#{MUTED}'>{contact}</font>", styles["Normal"]),
                title_block,
            ],
        ],
        colWidths=[50 * mm, 82 * mm, 38 * mm],
        style=[
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("LINEBELOW", (0, 0), (-1, -1), 2.2, colors.HexColor(f"#{RED}")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (2, 0), (2, 0), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ],
    )


def project_label(data, document):
    project = find_by_id(data.get("projects", []), document.get("projectId")) or {}
    code = project.get("code") or document.get("projectCode") or ""
    name = project.get("name") or document.get("projectName") or ""
    if code and name:
        return f"{code} - {name}"
    return code or name or "Not specified"


def document_client(data, document):
    client = find_by_id(data.get("clients", []), document.get("clientId")) or document.get("clientSnapshot") or {}
    return {
        "name": client.get("name", "Client"),
        "contact": client.get("contact", ""),
        "email": client.get("email", ""),
        "phone": client.get("phone", ""),
        "address": client.get("address", ""),
    }


def info_card(title, rows, styles):
    content = [[rich(f"<b>{escape(title)}</b>", styles["BoxTitle"])]]
    for label, value in rows:
        if value:
            content.append([rich(f"<font size='7' color='#{MUTED}'>{escape(label.upper())}</font><br/>{escape(str(value))}", styles["Normal"])])
    return Table(
        content,
        colWidths=[78 * mm],
        style=[
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(f"#{PAPER_TINT}")),
            ("BOX", (0, 0), (-1, -1), 0.45, colors.HexColor(f"#{BORDER}")),
            ("LINEBELOW", (0, 0), (0, 0), 0.8, colors.HexColor(f"#{RED}")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ],
    )


def document_message(kind, document, client, project):
    if kind == "invoice":
        return (
            f"Invoice {document.get('number', '')} for {client.get('name', 'the client')}."
            f" Please use this document number as the payment reference."
        )
    return (
        f"Quotation {document.get('number', '')} for {client.get('name', 'the client')}."
        f" This proposal covers the scope and pricing for {project}."
    )


def document_intro(kind, document, client, project, styles, data=None):
    message = document_message(kind, document, client, project)
    total_label = "Balance Due" if kind == "invoice" else "Proposal Value"
    action_label = "Payment requested" if kind == "invoice" else "Prepared for review"
    amount = max(0, document_total(document) - (invoice_paid(data or {}, document.get("id")) if kind == "invoice" else 0))
    return Table(
        [[
            rich(f"<font color='#{RED}'><b>{escape(action_label.upper())}</b></font><br/><b>{escape(message)}</b>", styles["Normal"]),
            rich(f"<font size='7' color='#{MUTED}'>{escape(total_label.upper())}</font><br/><font size='18' color='#{RED}'><b>{money(amount)}</b></font>", styles["Normal"]),
        ]],
        colWidths=[112 * mm, 58 * mm],
        style=[
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF7F7")),
            ("LINELEFT", (0, 0), (0, 0), 3.5, colors.HexColor(f"#{RED}")),
            ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor("#F4C9CC")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 12),
            ("RIGHTPADDING", (0, 0), (-1, -1), 12),
            ("TOPPADDING", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ],
    )


def totals_table(document, paid, styles):
    subtotal = sum(line_total(item) for item in document.get("items", []))
    discount = float(document.get("discount") or 0)
    tax = float(document.get("taxAmount") or max(0, subtotal - discount) * (float(document.get("taxRate") or 0) / 100))
    rows = [[p("Subtotal", styles["Normal"]), p(money(subtotal), styles["Normal"])]]
    if discount:
        rows.append([p("Discount", styles["Normal"]), p(f"({money(discount)})", styles["Normal"])])
    if tax:
        rows.append([p("VAT / Tax", styles["Normal"]), p(money(tax), styles["Normal"])])
    rows.extend([
        [p("Amount paid", styles["Normal"]), p(money(paid), styles["Normal"])],
        [p("Balance due", styles["Normal"]), p(money(max(0, document_total(document) - paid)), styles["Normal"])],
        [rich("<b>Total</b>", styles["Normal"]), rich(f"<font color='#{RED}'><b>{money(document_total(document))}</b></font>", styles["Normal"])],
    ])
    return Table(
        rows,
        colWidths=[42 * mm, 40 * mm],
        style=[
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, -1), (-1, -1), 12),
            ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor(f"#{RED}")),
            ("LINEABOVE", (0, -1), (-1, -1), 1.2, colors.HexColor(f"#{BLACK}")),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFF7F7")),
            ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor(f"#{BORDER}")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ],
    )


def statement_balance_table(balance, styles):
    rows = [
        [p("Current account balance", styles["Normal"]), rich(f"<font color='#{RED}'><b>{money(balance)}</b></font>", styles["Normal"])],
    ]
    return Table(
        rows,
        colWidths=[55 * mm, 40 * mm],
        style=[
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 12),
            ("BOX", (0, 0), (-1, -1), 0.45, colors.HexColor(f"#{BORDER}")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF7F7")),
            ("LINEABOVE", (0, 0), (-1, -1), 1.2, colors.HexColor(f"#{BLACK}")),
            ("LEFTPADDING", (0, 0), (-1, -1), 9),
            ("RIGHTPADDING", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ],
    )


def banking_box(company, styles):
    rows = [
        [rich("<b>Banking Details for EFT Payments</b>", styles["BoxTitle"])],
        [p(f"Bank: {company['bank']}", styles["Normal"])],
        [p(f"Account Holder: {company['accountHolder']}", styles["Normal"])],
        [p(f"Account Type: {company['accountType']}", styles["Normal"])],
        [p(f"Account Number: {company['accountNumber']}", styles["Normal"])],
        [p(f"Branch: {company['branchName']} | Branch Code: {company['branchCode']}", styles["Normal"])],
    ]
    return Table(rows, colWidths=[86 * mm], style=soft_box_style())


def document_party_panel(title, rows, styles, width=82):
    visible = [(label, value) for label, value in rows if value]
    if not visible:
        visible = [("Details", "Not specified")]
    content = [[rich(f"<font color='#{RED}'><b>{escape(title.upper())}</b></font>", styles["BoxTitle"])]]
    for label, value in visible:
        content.append([rich(f"<font size='7' color='#{MUTED}'>{escape(label.upper())}</font><br/><b>{escape(str(value))}</b>", styles["Normal"])])
    return Table(
        content,
        colWidths=[width * mm],
        style=[
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.45, colors.HexColor(f"#{BORDER}")),
            ("LINEABOVE", (0, 0), (-1, 0), 2, colors.HexColor(f"#{RED}")),
            ("LEFTPADDING", (0, 0), (-1, -1), 9),
            ("RIGHTPADDING", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ],
    )


def meta_panel(rows, styles, width=82):
    content = []
    for label, value in rows:
        content.append([p(label, styles["SmallMuted"]), rich(f"<b>{escape(str(value or ''))}</b>", styles["Normal"])])
    return Table(
        content,
        colWidths=[32 * mm, (width - 32) * mm],
        style=[
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(f"#{PAPER_TINT}")),
            ("BOX", (0, 0), (-1, -1), 0.45, colors.HexColor(f"#{BORDER}")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor(f"#{BORDER}")),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ],
    )


def notes_box(title, text, styles):
    return Table(
        [[rich(f"<b>{escape(title)}</b>", styles["BoxTitle"])], [p(text or "None", styles["SmallMuted"])]],
        colWidths=[78 * mm],
        style=soft_box_style(),
    )


def soft_box_style():
    return [
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(f"#{PAPER_TINT}")),
        ("BOX", (0, 0), (-1, -1), 0.45, colors.HexColor(f"#{BORDER}")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]


def build_document_pdf(data, kind, document):
    styles = stylesheet()
    company = company_from_data(data)
    client = document_client(data, document)
    project = project_label(data, document)
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    due_label = "Due date" if kind == "invoice" else "Valid until"
    due_value = document.get("dueDate") if kind == "invoice" else document.get("validUntil")
    paid = invoice_paid(data, document.get("id")) if kind == "invoice" else 0
    status = invoice_status(data, document) if kind == "invoice" else str(document.get("status") or "Draft").title()
    story = [
        brand_header(kind.title(), document.get("number", ""), styles, data),
        Spacer(1, 5 * mm),
        document_intro(kind, document, client, project, styles, data),
        Spacer(1, 7 * mm),
        Table(
            [[
                document_party_panel("Bill To", [
                    ("Name", client.get("name")),
                    ("Contact", client.get("contact")),
                    ("Phone", client.get("phone")),
                    ("Email", client.get("email")),
                    ("Address", client.get("address")),
                ], styles, 84),
                meta_panel([
                    ("Number", document.get("number")),
                    ("Date", document.get("date")),
                    (due_label, due_value),
                    ("Project", project),
                    ("Status", status),
                    ("Balance", money(max(0, document_total(document) - paid))),
                ], styles, 84),
            ]],
            colWidths=[84 * mm, 84 * mm],
            style=[
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ],
        ),
        Spacer(1, 8 * mm),
        rich("<b>Scope and Pricing</b>", styles["SectionTitle"]),
        Spacer(1, 1.5 * mm),
    ]
    rows = [[
        p("Description", styles["BoxTitle"]),
        p("Service", styles["BoxTitle"]),
        p("Qty", styles["BoxTitle"]),
        p("Rate", styles["BoxTitle"]),
        p("Amount", styles["BoxTitle"]),
    ]]
    for item in document.get("items", []):
        rows.append([
            p(item.get("description", ""), styles["Normal"]),
            p(service_name(data, item.get("serviceId") or document.get("serviceId")), styles["SmallMuted"]),
            p(item.get("qty", 0), styles["Normal"]),
            p(money(item.get("rate")), styles["Normal"]),
            p(money(line_total(item)), styles["Normal"]),
        ])
    if len(rows) == 1:
        rows.append([
            p("Service fee", styles["Normal"]),
            p(service_name(data, document.get("serviceId")), styles["SmallMuted"]),
            p("1", styles["Normal"]),
            p(money(document_total(document)), styles["Normal"]),
            p(money(document_total(document)), styles["Normal"]),
        ])
    table = Table(rows, colWidths=[70 * mm, 35 * mm, 16 * mm, 24 * mm, 25 * mm], repeatRows=1)
    table.setStyle(document_table_style())
    story.extend([
        table,
        Spacer(1, 7 * mm),
        Table(
            [[
                notes_box("Terms and Notes", f"{company['terms']}\n\n{document.get('notes') or company.get('notes')}", styles),
                totals_table(document, paid, styles),
            ]],
            colWidths=[84 * mm, 82 * mm],
            style=[("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0)],
        ),
        Spacer(1, 7 * mm),
        banking_box(company, styles),
        Spacer(1, 8 * mm),
        signature_table(company),
        Spacer(1, 5 * mm),
        rich(f"<font color='#{MUTED}'>{escape(company['subtitle'])}</font>", styles["Footer"]),
    ])
    pdf.build(story)
    return buffer.getvalue()


def document_table_style():
    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{BLACK}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{PAPER_TINT}")]),
            ("LINEBELOW", (0, 0), (-1, 0), 1.2, colors.HexColor(f"#{RED}")),
            ("LINEBELOW", (0, 1), (-1, -1), 0.3, colors.HexColor(f"#{BORDER}")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ]
    )


def signature_table(company=None):
    company = company or COMPANY
    return Table(
        [["Prepared by", "Approved / Client Signature"], [company.get("preparedBy", ""), company.get("approvedBy", "")]],
        colWidths=[80 * mm, 80 * mm],
        rowHeights=[8 * mm, 16 * mm],
        style=[
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{PAPER_TINT}")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(f"#{MUTED}")),
            ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor(f"#{BORDER}")),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor(f"#{BORDER}")),
            ("LINEBELOW", (0, 1), (-1, 1), 0.8, colors.HexColor(f"#{BLACK}")),
            ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ],
    )


def build_receipt_pdf(data, payment):
    styles = stylesheet()
    company = company_from_data(data)
    invoice = find_by_id(data.get("invoices", []), payment.get("invoiceId")) or {}
    client = document_client(data, invoice)
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm, topMargin=16 * mm, bottomMargin=16 * mm)
    allocation_rows = [[p("Applied To", styles["BoxTitle"]), p("Invoice Total", styles["BoxTitle"]), p("Paid This Receipt", styles["BoxTitle"]), p("Balance After", styles["BoxTitle"])]]
    invoice_total = document_total(invoice)
    paid_before = max(0, invoice_paid(data, invoice.get("id")) - float(payment.get("amount") or 0))
    balance_after = max(0, invoice_total - paid_before - float(payment.get("amount") or 0))
    allocation_rows.append([
        p(invoice.get("number", "Invoice"), styles["Normal"]),
        p(money(invoice_total), styles["Normal"]),
        p(money(payment.get("amount")), styles["Normal"]),
        p(money(balance_after), styles["Normal"]),
    ])
    allocation_table = Table(allocation_rows, colWidths=[62 * mm, 36 * mm, 38 * mm, 34 * mm], repeatRows=1)
    allocation_table.setStyle(document_table_style())
    story = [
        brand_header("Receipt", payment.get("receiptNumber", ""), styles, data),
        Spacer(1, 7 * mm),
        Table(
            [[
                Table(
                    [[rich("<font color='#5F6672'>AMOUNT RECEIVED</font>", styles["SmallMuted"])],
                     [rich(f"<font size='26' color='#{RED}'><b>{money(payment.get('amount'))}</b></font>", styles["Normal"])],
                     [p(f"Received from {client.get('name', 'Client')}", styles["Normal"])]],
                    colWidths=[70 * mm],
                    style=[
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF7F7")),
                        ("BOX", (0, 0), (-1, -1), 0.45, colors.HexColor("#F4C9CC")),
                        ("LINELEFT", (0, 0), (0, -1), 3, colors.HexColor(f"#{RED}")),
                        ("LEFTPADDING", (0, 0), (-1, -1), 12),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                        ("TOPPADDING", (0, 0), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ],
                ),
                meta_panel([
                    ("Receipt number", payment.get("receiptNumber")),
                    ("Payment date", payment.get("date")),
                    ("Invoice", invoice.get("number")),
                    ("Method", payment.get("method")),
                    ("Reference", payment.get("reference")),
                ], styles, 92),
            ]],
            colWidths=[74 * mm, 94 * mm],
            style=[("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0)],
        ),
        Spacer(1, 7 * mm),
        Table(
            [[
                document_party_panel("Received From", [("Client", client.get("name")), ("Phone", client.get("phone")), ("Email", client.get("email"))], styles, 84),
                notes_box("Receipt Note", f"Thank you for your payment. This receipt confirms funds recorded against invoice {invoice.get('number', '')}.", styles),
            ]],
            colWidths=[84 * mm, 84 * mm],
            style=[("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0)],
        ),
        Spacer(1, 7 * mm),
        rich("<b>Payment Allocation</b>", styles["SectionTitle"]),
        Spacer(1, 1.5 * mm),
        allocation_table,
        Spacer(1, 7 * mm),
        banking_box(company, styles),
        Spacer(1, 8 * mm),
        signature_table(company),
        Spacer(1, 5 * mm),
        rich(f"<font color='#{MUTED}'>{escape(company['subtitle'])}</font>", styles["Footer"]),
    ]
    pdf.build(story)
    return buffer.getvalue()


def statement_rows(data, statement_type, party_id):
    if statement_type == "supplier":
        party = find_by_id(data.get("suppliers", []), party_id) or {}
        balance = float(party.get("openingBalance") or 0)
        rows = []
        if balance:
            rows.append([party.get("createdAt", ""), "Opening balance", "", money(0), money(balance), money(balance)])
        for bill in data.get("supplierBills", []):
            if bill.get("supplierId") == party_id and str(bill.get("status", "")).lower() not in ["voided", "cancelled", "archived", "deleted"]:
                amount = float(bill.get("amount") or document_total(bill))
                balance += amount
                rows.append([bill.get("date", ""), "Supplier bill", bill.get("number", ""), money(amount), money(0), money(balance)])
        for payment in data.get("supplierPayments", []):
            if payment.get("supplierId") == party_id and str(payment.get("status", "")).lower() not in ["voided", "cancelled", "archived", "deleted"]:
                amount = float(payment.get("amount") or 0)
                balance -= amount
                rows.append([payment.get("date", ""), "Supplier payment", payment.get("reference", ""), money(0), money(amount), money(balance)])
        return party, rows, balance

    party = find_by_id(data.get("clients", []), party_id) or {}
    balance = float(party.get("openingBalance") or 0)
    rows = []
    if balance:
        rows.append([party.get("createdAt", ""), "Opening balance", "", money(balance), money(0), money(balance)])
    for invoice in data.get("invoices", []):
        if invoice.get("clientId") == party_id and str(invoice.get("status", "")).lower() not in ["draft", "pending-review", "rejected", "voided", "cancelled", "archived", "deleted"]:
            amount = document_total(invoice)
            balance += amount
            rows.append([invoice.get("date", ""), "Invoice", invoice.get("number", ""), money(amount), money(0), money(balance)])
    for payment in data.get("payments", []):
        if payment.get("clientId") == party_id and str(payment.get("status", "")).lower() not in ["voided", "cancelled", "archived", "deleted"]:
            amount = float(payment.get("amount") or 0)
            balance -= amount
            rows.append([payment.get("date", ""), "Receipt", payment.get("receiptNumber", ""), money(0), money(amount), money(balance)])
    return party, rows, balance


def build_statement_pdf(data, statement_type, party_id):
    styles = stylesheet()
    company = company_from_data(data)
    party, rows, balance = statement_rows(data, statement_type, party_id)
    title = "Supplier Statement" if statement_type == "supplier" else "Statement of Account"
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    table_rows = [[p("Date", styles["BoxTitle"]), p("Type", styles["BoxTitle"]), p("Number", styles["BoxTitle"]), p("Debit", styles["BoxTitle"]), p("Credit", styles["BoxTitle"]), p("Balance", styles["BoxTitle"])]]
    for row in rows:
        table_rows.append([p(cell, styles["Normal"]) for cell in row])
    if len(table_rows) == 1:
        table_rows.append([p("No statement transactions", styles["Normal"]), "", "", "", "", ""])
    table = Table(table_rows, colWidths=[24 * mm, 34 * mm, 30 * mm, 27 * mm, 27 * mm, 28 * mm], repeatRows=1)
    table.setStyle(document_table_style())
    story = [
        brand_header(title, datetime.now().strftime("%Y-%m-%d"), styles, data),
        Spacer(1, 7 * mm),
        Table(
            [[
                document_party_panel("Account Holder", [
                    ("Name", party.get("name")),
                    ("Contact", party.get("contact")),
                    ("Phone", party.get("phone")),
                    ("Email", party.get("email")),
                    ("Address", party.get("address")),
                ], styles, 84),
                meta_panel([
                    ("Statement date", datetime.now().strftime("%Y-%m-%d")),
                    ("Account type", "Creditor" if statement_type == "supplier" else "Debtor"),
                    ("Balance", money(balance)),
                ], styles, 84),
            ]],
            colWidths=[84 * mm, 84 * mm],
            style=[("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0)],
        ),
        Spacer(1, 8 * mm),
        rich("<b>Account Activity</b>", styles["SectionTitle"]),
        table,
        Spacer(1, 7 * mm),
        statement_balance_table(balance, styles),
        Spacer(1, 7 * mm),
        banking_box(company, styles),
        Spacer(1, 8 * mm),
        signature_table(company),
        Spacer(1, 5 * mm),
        rich(f"<font color='#{MUTED}'>{escape(company['subtitle'])}</font>", styles["Footer"]),
    ]
    pdf.build(story)
    return buffer.getvalue()


def report_table(headers, rows, page_width_mm=250):
    clean_headers = [p(header, stylesheet()["BoxTitle"]) for header in headers]
    clean_rows = []
    styles = stylesheet()
    for row in rows:
        clean_rows.append([p(cell, styles["Normal"]) for cell in row])
    table_rows = [clean_headers] + clean_rows if headers else clean_rows
    col_count = max(len(table_rows[0]), 1) if table_rows else 1
    widths = [page_width_mm / col_count * mm for _ in range(col_count)]
    table = Table(table_rows or [[p("No records", styles["Normal"])]], colWidths=widths, repeatRows=1)
    table.setStyle(document_table_style())
    return table


def metric_cards(metrics, styles):
    cards = []
    for label, value in metrics:
        cards.append(rich(f"<font size='7' color='#{MUTED}'>{escape(label.upper())}</font><br/><font size='14' color='#{BLACK}'><b>{escape(str(value))}</b></font>", styles["Normal"]))
    return Table(
        [cards],
        colWidths=[58 * mm for _ in cards],
        style=[
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF7F7")),
            ("BOX", (0, 0), (-1, -1), 0.45, colors.HexColor("#F4C9CC")),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#F4C9CC")),
            ("LEFTPADDING", (0, 0), (-1, -1), 9),
            ("RIGHTPADDING", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ],
    )


def build_report_pdf(data):
    styles = stylesheet()
    company = company_from_data(data)
    summary = current_month_summary(data)
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=16 * mm, rightMargin=16 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    invoices = []
    for invoice in data.get("invoices", []):
        invoices.append([invoice.get("number", ""), client_name(data, invoice.get("clientId")), money(document_total(invoice)), money(invoice_paid(data, invoice.get("id"))), invoice_status(data, invoice)])
    story = [
        brand_header("Monthly Financial Report", summary["month"], styles, data),
        Spacer(1, 6 * mm),
        rich("<b>Executive Summary</b>", styles["DocTitle"]),
        metric_cards([
            ("Invoice value issued", money(summary["invoiced"])),
            ("Cash received", money(summary["received"])),
            ("Expenses recorded", money(summary["expenses"])),
            ("Net cash position", money(summary["net"])),
        ], styles),
        Spacer(1, 7 * mm),
        notes_box("Management Note", "This report summarizes the current month position for quick management review. It is intended for decision-making and follow-up, not only data storage.", styles),
        Spacer(1, 7 * mm),
        rich("<b>Invoice Payment Position</b>", styles["DocTitle"]),
        report_table(["Invoice", "Client", "Total", "Paid", "Status"], invoices, 250),
        Spacer(1, 10 * mm),
        signature_table(company),
        Spacer(1, 4 * mm),
        rich(f"<font color='#{MUTED}'>{escape(company['subtitle'])}</font>", styles["Footer"]),
    ]
    pdf.build(story)
    return buffer.getvalue()


def build_generic_report_pdf(report, data=None):
    styles = stylesheet()
    company = company_from_data(data or {})
    title = report.get("title", "Accounting Report")
    headers = report.get("headers", [])
    rows = report.get("rows", [])
    buffer = BytesIO()
    wide = len(headers or []) > 4
    page_size = landscape(A4) if wide else A4
    page_width = 250 if wide else 170
    pdf = SimpleDocTemplate(buffer, pagesize=page_size, leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    story = [
        brand_header(title, datetime.now().strftime("%Y-%m-%d"), styles, data),
        Spacer(1, 6 * mm),
        notes_box("Report Purpose", "Prepared from Civil-Gineer Masta business records. Use this report for review, reconciliation, approvals, and management action.", styles),
        Spacer(1, 6 * mm),
        report_table(headers, rows, page_width),
        Spacer(1, 9 * mm),
        signature_table(company),
        Spacer(1, 4 * mm),
        rich(f"<font color='#{MUTED}'>{escape(company['subtitle'])}</font>", styles["Footer"]),
    ]
    pdf.build(story)
    return buffer.getvalue()


def build_workbook(data):
    wb = Workbook()
    wb.remove(wb.active)
    add_summary_sheet(wb, data)
    add_table_sheet(wb, "Clients", ["Code", "Name", "Contact", "Email", "Phone", "Address", "Tax/VAT", "Opening Balance", "Created"], [
        [c.get("code", ""), c.get("name", ""), c.get("contact", ""), c.get("email", ""), c.get("phone", ""), c.get("address", ""), c.get("taxId", ""), c.get("openingBalance", 0), c.get("createdAt", "")]
        for c in data.get("clients", [])
    ], money_cols=[8], date_cols=[9])
    add_table_sheet(wb, "Projects", ["Code", "Project", "Client", "Service", "Location", "Status", "Start Date", "Expected Completion", "Notes"], [
        [p.get("code", ""), p.get("name", ""), client_name(data, p.get("clientId")), service_name(data, p.get("serviceId")), p.get("location", ""), p.get("status", ""), p.get("startDate", ""), p.get("expectedCompletionDate", ""), p.get("notes", "")]
        for p in data.get("projects", [])
    ], date_cols=[7, 8], status_cols=[6])
    add_table_sheet(wb, "Services", ["Service", "Income Account", "Cost Account", "Description"], [
        [s.get("name", ""), s.get("incomeAccountId", ""), s.get("costAccountId", ""), s.get("description", "")]
        for s in data.get("services", [])
    ])
    add_table_sheet(wb, "Quotations", ["Number", "Client", "Project", "Main Service", "Date", "Valid Until", "Status", "Subtotal", "Discount", "VAT/Tax", "Total", "Notes"], [
        [q.get("number", ""), quotation_client_name(data, q), workbook_project_label(data, q), service_name(data, q.get("serviceId")), q.get("date", ""), q.get("validUntil", ""), q.get("status", ""), sum(line_total(item) for item in q.get("items", [])), q.get("discount", 0), document_total(q) - max(0, sum(line_total(item) for item in q.get("items", [])) - float(q.get("discount") or 0)), document_total(q), q.get("notes", "")]
        for q in data.get("quotations", [])
    ], money_cols=[8, 9, 10, 11], date_cols=[5, 6], status_cols=[7])
    add_table_sheet(wb, "Quotation Items", ["Quotation", "Client", "Project", "Service Category", "Description", "Qty", "Rate", "Line Total"], [
        [q.get("number", ""), quotation_client_name(data, q), workbook_project_label(data, q), service_name(data, item.get("serviceId") or q.get("serviceId")), item.get("description", ""), item.get("qty", 0), item.get("rate", 0), line_total(item)]
        for q in data.get("quotations", []) for item in q.get("items", [])
    ], money_cols=[7, 8])
    add_table_sheet(wb, "Invoices", ["Number", "Client", "Project", "Main Service", "Date", "Due Date", "Subtotal", "Discount", "VAT/Tax", "Total", "Paid", "Outstanding", "Status", "Notes"], [
        [i.get("number", ""), client_name(data, i.get("clientId")), workbook_project_label(data, i), service_name(data, i.get("serviceId")), i.get("date", ""), i.get("dueDate", ""), sum(line_total(item) for item in i.get("items", [])), i.get("discount", 0), document_total(i) - max(0, sum(line_total(item) for item in i.get("items", [])) - float(i.get("discount") or 0)), document_total(i), invoice_paid(data, i.get("id")), document_total(i) - invoice_paid(data, i.get("id")), invoice_status(data, i), i.get("notes", "")]
        for i in data.get("invoices", [])
    ], money_cols=[7, 8, 9, 10, 11, 12], date_cols=[5, 6], status_cols=[13])
    add_table_sheet(wb, "Invoice Items", ["Invoice", "Client", "Project", "Service Category", "Description", "Qty", "Rate", "Line Total"], [
        [i.get("number", ""), client_name(data, i.get("clientId")), workbook_project_label(data, i), service_name(data, item.get("serviceId") or i.get("serviceId")), item.get("description", ""), item.get("qty", 0), item.get("rate", 0), line_total(item)]
        for i in data.get("invoices", []) for item in i.get("items", [])
    ], money_cols=[7, 8])
    add_table_sheet(wb, "Payments", ["Date", "Receipt", "Invoice", "Client", "Project", "Service", "Bank/Cash Account", "Amount", "Method", "Reference", "Status"], [
        [p.get("date", ""), p.get("receiptNumber", ""), invoice_number(data, p.get("invoiceId")), payment_client(data, p), workbook_project_label(data, p), service_name(data, p.get("serviceId")), account_name(data, p.get("bankAccountId")), p.get("amount", 0), p.get("method", ""), p.get("reference", ""), p.get("status", "paid")]
        for p in data.get("payments", [])
    ], money_cols=[8], date_cols=[1], status_cols=[11])
    add_table_sheet(wb, "Receipts", ["Receipt", "Date", "Client", "Invoice", "Project", "Service", "Amount", "Method", "Reference", "Status"], [
        [p.get("receiptNumber", ""), p.get("date", ""), payment_client(data, p), invoice_number(data, p.get("invoiceId")), workbook_project_label(data, p), service_name(data, p.get("serviceId")), p.get("amount", 0), p.get("method", ""), p.get("reference", ""), p.get("status", "paid")]
        for p in data.get("payments", [])
    ], money_cols=[7], date_cols=[2], status_cols=[10])
    add_table_sheet(wb, "Expenses", ["Date", "Category", "Vendor", "Project", "Service", "Description", "Amount", "Payment Method", "Status"], [
        [e.get("date", ""), e.get("category", ""), e.get("vendor", ""), workbook_project_label(data, e), service_name(data, e.get("serviceId")), e.get("description", ""), e.get("amount", 0), e.get("paymentMethod", ""), e.get("status", "paid")]
        for e in data.get("expenses", [])
    ], money_cols=[7], date_cols=[1], status_cols=[9])
    add_table_sheet(wb, "Suppliers", ["Name", "Contact", "Email", "Phone", "Address", "Opening Balance", "Status"], [
        [s.get("name", ""), s.get("contact", ""), s.get("email", ""), s.get("phone", ""), s.get("address", ""), s.get("openingBalance", 0), s.get("status", "active")]
        for s in data.get("suppliers", [])
    ], money_cols=[6], status_cols=[7])
    add_table_sheet(wb, "Supplier Bills", ["Number", "Supplier", "Project", "Service", "Date", "Due Date", "Description", "Amount", "Status"], [
        [b.get("number", ""), supplier_name(data, b.get("supplierId")), workbook_project_label(data, b), service_name(data, b.get("serviceId")), b.get("date", ""), b.get("dueDate", ""), b.get("description", ""), b.get("amount", 0), b.get("status", "issued")]
        for b in data.get("supplierBills", [])
    ], money_cols=[8], date_cols=[5, 6], status_cols=[9])
    add_table_sheet(wb, "Supplier Payments", ["Date", "Supplier", "Bill", "Amount", "Reference", "Status"], [
        [p.get("date", ""), supplier_name(data, p.get("supplierId")), supplier_bill_number(data, p.get("billId")), p.get("amount", 0), p.get("reference", ""), p.get("status", "paid")]
        for p in data.get("supplierPayments", [])
    ], money_cols=[4], date_cols=[1], status_cols=[6])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_generic_report_workbook(report):
    wb = Workbook()
    wb.remove(wb.active)
    title = report.get("title", "Accounting Report")
    headers = report.get("headers", [])
    rows = report.get("rows", [])
    money_cols = [index + 1 for index, header in enumerate(headers) if any(term in str(header).lower() for term in ["amount", "total", "debit", "credit", "balance", "income", "expense", "cash", "profit", "loss", "paid", "owed", "outstanding"])]
    add_table_sheet(wb, "Report", headers, rows, money_cols=money_cols, display_title=title, subtitle="Prepared from Civil-Gineer Masta business records")
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def invoice_number(data, invoice_id):
    invoice = find_by_id(data.get("invoices", []), invoice_id) or {}
    return invoice.get("number", "")


def payment_client(data, payment):
    invoice = find_by_id(data.get("invoices", []), payment.get("invoiceId")) or {}
    return client_name(data, invoice.get("clientId"))


def add_summary_sheet(wb, data):
    ws = wb.create_sheet("Monthly Summary")
    summary = current_month_summary(data)
    rows = [
        ["Civil-Gineer Masta", "", "", ""],
        ["Excel Accounting Export", f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}", "", ""],
        ["Current month", summary["month"], "", ""],
        ["", "", "", ""],
        ["Metric", "Amount", "Formula / Notes", "Management meaning"],
        ["Invoice value issued", "=SUM(Invoices!J5:J1048576)", "Pulls from invoice total column", "Work billed to clients this month"],
        ["Cash received", "=SUM(Payments!H5:H1048576)", "Pulls from payment amount column", "Money received and receipted"],
        ["Expenses recorded", "=SUM(Expenses!G5:G1048576)", "Pulls from expense amount column", "Cash costs captured"],
        ["Net cash position", "=B7-B8", "Cash received minus expenses", "Positive means cash received is higher than expenses"],
        ["Outstanding invoices", "=SUM(Invoices!L5:L1048576)", "Pulls from invoice outstanding column", "Client money still to collect"],
        ["Supplier bills payable", "=SUM('Supplier Bills'!H5:H1048576)", "Pulls from supplier bills amount column", "Supplier obligations recorded"],
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws, money_cols=[2], title_rows=4, header_row=5)
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")
    ws["A1"].font = Font(bold=True, size=18, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=BLACK)
    ws["A2"].font = Font(bold=True, size=12, color=RED)
    ws["A3"].font = Font(bold=True, color=BLACK)
    ws.sheet_view.showGridLines = False


def add_table_sheet(wb, title, headers, rows, money_cols=None, date_cols=None, status_cols=None, subtitle=None, display_title=None):
    ws = wb.create_sheet(title[:31])
    last_col = max(len(headers), 1)
    ws.append([display_title or title] + [""] * (last_col - 1))
    ws.append([subtitle or f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"] + [""] * (last_col - 1))
    ws.append([""] * last_col)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    header_row = 4
    first_data_row = 5
    if rows:
        total_row = ws.max_row + 1
        ws.cell(total_row, 1, "Totals")
        for col in money_cols or []:
            letter = get_column_letter(col)
            ws.cell(total_row, col, f"=SUM({letter}{first_data_row}:{letter}{total_row - 1})")
        ws.cell(total_row, 1).font = Font(bold=True, color=BLACK)
    else:
        ws.append(["No records captured yet"] + [""] * (last_col - 1))
    style_sheet(ws, money_cols=money_cols or [], date_cols=date_cols or [], status_cols=status_cols or [], title_rows=3, header_row=header_row)
    data_end = ws.max_row
    if data_end >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{data_end}"
        if rows:
            table_ref = f"A{header_row}:{get_column_letter(last_col)}{data_end}"
            table_name = safe_table_name(title)
            table = ExcelTable(displayName=table_name, ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)


def safe_table_name(title):
    clean = "".join(char for char in title.title().replace(" ", "") if char.isalnum())
    return f"{clean[:24] or 'Export'}Table"


def style_sheet(ws, money_cols=None, date_cols=None, status_cols=None, title_rows=0, header_row=1):
    header_fill = PatternFill("solid", fgColor=RED)
    title_fill = PatternFill("solid", fgColor=BLACK)
    section_fill = PatternFill("solid", fgColor="FFF4F4")
    light_fill = PatternFill("solid", fgColor=PAPER_TINT)
    thin = Side(style="thin", color="D8DDE4")
    medium = Side(style="medium", color=BLACK)
    max_col = max(ws.max_column, 1)
    if title_rows:
        for row_index in range(1, title_rows + 1):
            if row_index <= 2:
                ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=max_col)
            for cell in ws[row_index]:
                cell.fill = title_fill if row_index == 1 else section_fill
                cell.font = Font(bold=True, color=WHITE if row_index == 1 else BLACK, size=15 if row_index == 1 else 10)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = Border(bottom=medium if row_index == 1 else thin)
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 20
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row > header_row and cell.row % 2 == 0:
                cell.fill = light_fill
    for cell in ws[header_row]:
        cell.fill = header_fill if ws.title != "Monthly Summary" else title_fill
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col in money_cols or []:
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=header_row + 1):
            for item in cell:
                item.number_format = '"BWP" #,##0.00'
                item.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    for col in date_cols or []:
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=header_row + 1):
            for item in cell:
                item.number_format = "yyyy-mm-dd"
    for col in status_cols or []:
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=header_row + 1):
            for item in cell:
                style_status_cell(item)
    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.showGridLines = False
    for index, column in enumerate(ws.columns, start=1):
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[get_column_letter(index)].width = min(max(max_len + 2, 12), 38)
    ws.row_dimensions[header_row].height = 22


def style_status_cell(cell):
    value = str(cell.value or "").lower()
    if not value:
        return
    if any(term in value for term in ["paid", "approved", "issued", "accepted", "active"]):
        cell.fill = PatternFill("solid", fgColor="E8F6EE")
        cell.font = Font(color="16803C", bold=True)
    elif any(term in value for term in ["partial", "sent", "pending", "draft"]):
        cell.fill = PatternFill("solid", fgColor="FFF4DD")
        cell.font = Font(color="A66500", bold=True)
    elif any(term in value for term in ["overdue", "rejected", "voided", "cancelled"]):
        cell.fill = PatternFill("solid", fgColor="FFF1F1")
        cell.font = Font(color=RED, bold=True)
    elif "archived" in value:
        cell.fill = PatternFill("solid", fgColor="E9EDF2")
        cell.font = Font(color=MUTED, bold=True)


class ExportHandler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(204)
        self._cors()
        self.end_headers()

    def do_GET(self):
        if urlparse(self.path).path == "/health":
            self._send_json({"ok": True, "service": "CGM export backend"})
            return
        self.send_error(404)

    def do_POST(self):
        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length) or b"{}")
            data = payload.get("data", {})
            path = urlparse(self.path).path
            if path == "/api/export/pdf":
                content = self._pdf(payload, data)
                self._send_file(content, "application/pdf", payload.get("filename") or "cgm-export.pdf")
                return
            if path == "/api/export/excel":
                content = build_generic_report_workbook(payload["report"]) if payload.get("report") else build_workbook(data)
                self._send_file(content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", payload.get("filename") or "cgm-accounting-export.xlsx")
                return
            self.send_error(404)
        except Exception as exc:
            self._send_json({"ok": False, "error": str(exc)}, status=500)

    def _pdf(self, payload, data):
        kind = payload.get("kind")
        item_id = payload.get("id")
        if kind == "quotation":
            document = find_by_id(data.get("quotations", []), item_id)
            return build_document_pdf(data, "quotation", document)
        if kind == "invoice":
            document = find_by_id(data.get("invoices", []), item_id)
            return build_document_pdf(data, "invoice", document)
        if kind == "receipt":
            payment = find_by_id(data.get("payments", []), item_id)
            return build_receipt_pdf(data, payment)
        if kind == "statement":
            return build_statement_pdf(data, payload.get("statementType") or "client", item_id)
        if kind == "financial-report":
            return build_report_pdf(data)
        if kind == "generic-report":
            return build_generic_report_pdf(payload.get("report", {}), data)
        raise ValueError("Unsupported PDF export kind")

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def _send_file(self, content, content_type, filename):
        self.send_response(200)
        self._cors()
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def _send_json(self, payload, status=200):
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self._cors()
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format, *args):
        return


if __name__ == "__main__":
    print(f"CGM export backend running at http://{HOST}:{PORT}")
    ThreadingHTTPServer((HOST, PORT), ExportHandler).serve_forever()
